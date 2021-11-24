import json
from datetime import datetime
from io import BytesIO

import xlsxwriter
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView
from django.views.generic.base import View
from openpyxl import load_workbook

from core.pos.forms import ProductForm, Product, Category
from core.security.mixins import PermissionMixin, ModuleMixin


class ProductListView(PermissionMixin, TemplateView):
    template_name = 'scm/product/list.html'
    permission_required = 'view_product'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search':
                data = []
                for p in Product.objects.filter():
                    data.append(p.toJSON())
            elif action == 'upload_excel':
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]
                    for row in range(2, excel.max_row + 1):
                        product = Product()
                        id = int(excel.cell(row=row, column=1).value)
                        if Product.objects.filter(id=id).exists():
                            product = Product.objects.get(pk=id)
                        product.name = excel.cell(row=row, column=2).value
                        name = excel.cell(row=row, column=3).value
                        if not Category.objects.filter(name=name).exists():
                            category = Category()
                            category.name = name
                            category.save()
                        else:
                            category = Category.objects.get(name=name)
                        product.category_id = category.id
                        product.price = float(excel.cell(row=row, column=4).value)
                        product.pvp = float(excel.cell(row=row, column=5).value)
                        product.stock = int(excel.cell(row=row, column=6).value)
                        product.save()
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_url'] = reverse_lazy('product_create')
        context['title'] = 'Listado de Productos'
        return context


class ProductCreateView(PermissionMixin, CreateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'add_product'

    def validate_data(self):
        data = {'valid': True}
        try:
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                product = Product()
                product.name = request.POST['name']
                product.category_id = request.POST['category']
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.save()
            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Nuevo registro de un Producto'
        context['action'] = 'add'
        return context


class ProductUpdateView(PermissionMixin, UpdateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'change_product'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def validate_data(self):
        data = {'valid': True}
        try:
            id = self.get_object().id
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category).exclude(id=id):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                product = self.object
                product.name = request.POST['name']
                product.category_id = request.POST['category']
                if 'image-clear' in request.POST:
                    product.remove_image()
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.save()
            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Edición de un Producto'
        context['action'] = 'edit'
        return context


class ProductDeleteView(PermissionMixin, DeleteView):
    model = Product
    template_name = 'scm/product/delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'delete_product'

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.get_object().delete()
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class ProductStockAdjustmentView(ModuleMixin, TemplateView):
    template_name = 'scm/product/stock_adjustment.html'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_products':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                search = Product.objects.filter(category__inventoried=True).exclude(id__in=ids).order_by('name')
                if len(term):
                    search = search.filter(name__icontains=term)
                    search = search[0:10]
                for p in search:
                    item = p.toJSON()
                    item['value'] = '{} / {}'.format(p.name, p.category.name)
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for p in json.loads(request.POST['products']):
                        product = Product.objects.get(pk=p['id'])
                        product.stock = int(p['newstock'])
                        product.save()
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ajuste de Stock de Productos'
        return context


class ProductExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                'Código': 15,
                'Producto': 75,
                'Categoría': 20,
                'Precio de Compra': 20,
                'Precio de Venta': 20,
                'Stock': 10,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('productos')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=0, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            for product in Product.objects.filter().order_by('id'):
                worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 1, product.name, row_format)
                worksheet.write(row, 2, product.category.name, row_format)
                worksheet.write(row, 3, format(product.price, '.2f'), row_format)
                worksheet.write(row, 4, format(product.pvp, '.2f'), row_format)
                worksheet.write(row, 5, product.stock, row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="PRODUCTOS_{}.xlsx"'.format(
                datetime.now().date().strftime('%d_%m_%Y'))
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('product_list'))
